"""
data_exporter.py

Three modes:
  python data_exporter.py export          → dump all data to exports/ folder (JSON + Excel)
  python data_exporter.py serve           → local API server  (friend on same Wi-Fi)
  python data_exporter.py tunnel          → public HTTPS URL  (friend anywhere in the world)
  python data_exporter.py serve  --port=8080
  python data_exporter.py tunnel --port=8080

Tunnel mode uses ngrok — free, no sign-up needed for basic use.
Your friend gets a URL like:  https://abc123.ngrok-free.app/api/all
"""

import json
import os
import sys
import socket
from datetime import datetime, date, time

from db import fetch_all

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "exports")


# ── JSON helper (handles datetime / date / time objects) ──────────────────────
class _Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, time):
            return str(obj)
        return super().default(obj)


def _dump(obj):
    return json.dumps(obj, cls=_Encoder, ensure_ascii=False, indent=2)


# ── Fetch all tables ──────────────────────────────────────────────────────────
def fetch_schedule():
    return fetch_all("""
        SELECT
            sch.schedule_id,
            c.course_na      AS course_name,
            c.credit_hours,
            i.instructor_name,
            r.building       AS room_name,
            r.capacity,
            ts.day,
            ts.s_time        AS start_time,
            ts.e_time        AS end_time,
            sch.course_id,
            sch.instructor_id,
            sch.room_id,
            sch.time_id
        FROM schedule sch
        LEFT JOIN course      c  ON sch.course_id    = c.course_id
        LEFT JOIN instructor  i  ON sch.instructor_id = i.instructor_id
        LEFT JOIN room        r  ON sch.room_id       = r.room_id
        LEFT JOIN time_slot   ts ON sch.time_id       = ts.time_id
        ORDER BY sch.schedule_id
    """)


def fetch_students():
    return fetch_all("""
        SELECT
            s.std_id,
            s.std_na   AS student_name,
            s.major_id,
            m.major_na AS major_name
        FROM std s
        LEFT JOIN major m ON s.major_id = m.major_id
        ORDER BY s.std_id
    """)


def fetch_courses():
    return fetch_all("""
        SELECT
            c.course_id,
            c.course_na  AS course_name,
            c.credit_hours,
            c.type       AS course_type,
            c.major_id,
            m.major_na   AS major_name,
            c.prereq_id
        FROM course c
        LEFT JOIN major m ON c.major_id = m.major_id
        ORDER BY c.course_id
    """)


def fetch_instructors():
    return fetch_all("""
        SELECT
            instructor_id,
            instructor_name,
            ISNULL(spec, '')        AS specialization,
            ISNULL(degree_type, '') AS degree_type
        FROM instructor
        ORDER BY instructor_id
    """)


def fetch_rooms():
    return fetch_all("""
        SELECT
            room_id,
            building AS room_name,
            ISNULL(capacity, 40) AS capacity,
            ISNULL(type, 'Lecture') AS room_type
        FROM room
        ORDER BY room_id
    """)


def fetch_all_data():
    return {
        "exported_at": datetime.now().isoformat(),
        "schedule":    fetch_schedule(),
        "students":    fetch_students(),
        "courses":     fetch_courses(),
        "instructors": fetch_instructors(),
        "rooms":       fetch_rooms(),
    }


# ── FILE EXPORT ───────────────────────────────────────────────────────────────
def export_to_files():
    os.makedirs(EXPORT_DIR, exist_ok=True)

    data = fetch_all_data()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── JSON files ────────────────────────────────────────────────────────────
    for key in ("schedule", "students", "courses", "instructors", "rooms"):
        path = os.path.join(EXPORT_DIR, f"{key}.json")
        with open(path, "w", encoding="utf-8") as f:
            f.write(_dump(data[key]))
        print(f"[OK] {path}")

    # ── Full bundle JSON ──────────────────────────────────────────────────────
    bundle_path = os.path.join(EXPORT_DIR, f"full_export_{timestamp}.json")
    with open(bundle_path, "w", encoding="utf-8") as f:
        f.write(_dump(data))
    print(f"[OK] {bundle_path}")

    # ── Excel workbook ────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="2E75B6")
        center_align = Alignment(horizontal="center", vertical="center")

        for key in ("schedule", "students", "courses", "instructors", "rooms"):
            rows = data[key]
            if not rows:
                continue

            ws = wb.create_sheet(title=key.capitalize())
            headers = list(rows[0].keys())

            # header row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = center_align

            # data rows
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, key_name in enumerate(headers, start=1):
                    val = row.get(key_name)
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()
                    elif isinstance(val, time):
                        val = str(val)
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        excel_path = os.path.join(EXPORT_DIR, f"schedule_export_{timestamp}.xlsx")
        wb.save(excel_path)
        print(f"[OK] {excel_path}")

    except ImportError:
        print("[SKIP] openpyxl not installed — Excel export skipped. Run: pip install openpyxl")

    print(f"\nAll files saved to: {EXPORT_DIR}")
    print("Share the exports/ folder with your friend via USB, email, or cloud storage.")


# ── LOCAL API SERVER ──────────────────────────────────────────────────────────
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def run_server(port=5050):
    try:
        from flask import Flask, jsonify, request
        from flask_cors import CORS
    except ImportError:
        print("Flask not found. Installing...")
        os.system(f"{sys.executable} -m pip install flask flask-cors -q")
        from flask import Flask, jsonify, request
        from flask_cors import CORS

    app = Flask(__name__)
    CORS(app)  # allow cross-origin so friend can call from browser / any client

    def _resp(data):
        return app.response_class(
            response=_dump(data),
            status=200,
            mimetype="application/json"
        )

    @app.route("/")
    def index():
        local_ip = get_local_ip()
        return _resp({
            "service": "SSG Data Exporter",
            "endpoints": {
                "GET /api/schedule":     "Full schedule with rooms & instructors",
                "GET /api/students":     "All students",
                "GET /api/courses":      "All courses",
                "GET /api/instructors":  "All instructors",
                "GET /api/rooms":        "All rooms",
                "GET /api/all":          "Everything in one response",
            },
            "your_ip": local_ip,
            "port": port,
            "friend_url": f"http://{local_ip}:{port}/api/all",
        })

    @app.route("/api/schedule")
    def api_schedule():
        return _resp(fetch_schedule())

    @app.route("/api/students")
    def api_students():
        return _resp(fetch_students())

    @app.route("/api/courses")
    def api_courses():
        return _resp(fetch_courses())

    @app.route("/api/instructors")
    def api_instructors():
        return _resp(fetch_instructors())

    @app.route("/api/rooms")
    def api_rooms():
        return _resp(fetch_rooms())

    @app.route("/api/all")
    def api_all():
        return _resp(fetch_all_data())

    # ── Print connection info ──────────────────────────────────────────────────
    local_ip = get_local_ip()
    print("=" * 60)
    print("  SSG Data Exporter — API Server")
    print("=" * 60)
    print(f"  Local:        http://127.0.0.1:{port}/")
    print(f"  LAN (friend): http://{local_ip}:{port}/api/all")
    print()
    print("  Both laptops must be on the SAME Wi-Fi.")
    print("  For remote access run:  python data_exporter.py tunnel")
    print("=" * 60)

    return app, port


def _start_flask(app, port):
    app.run(host="0.0.0.0", port=port, debug=False)


# ── NGROK TUNNEL ──────────────────────────────────────────────────────────────
NGROK_TOKEN = "3CtmSPWzjjNy35noSp4nZtmwKVS_7z2VysvRef9LSXNL1ZmLr"


def run_tunnel(port=5050):
    try:
        from pyngrok import ngrok, conf
    except ImportError:
        print("pyngrok not found. Installing...")
        os.system(f"{sys.executable} -m pip install pyngrok -q")
        from pyngrok import ngrok, conf

    # set token directly so config file format doesn't matter
    ngrok.set_auth_token(NGROK_TOKEN)

    app, port = run_server(port)

    import threading

    # start Flask in a background thread so ngrok can run in main thread
    t = threading.Thread(target=_start_flask, args=(app, port), daemon=True)
    t.start()

    # open the tunnel
    tunnel = ngrok.connect(port, "http")
    public_url = tunnel.public_url

    print()
    print("=" * 60)
    print("  TUNNEL ACTIVE — friend can access from anywhere")
    print("=" * 60)
    print(f"  Public URL:  {public_url}/api/all")
    print()
    print("  Send this link to your friend:")
    print(f"    {public_url}/api/all          ← everything")
    print(f"    {public_url}/api/schedule     ← schedule only")
    print(f"    {public_url}/api/students     ← students only")
    print()
    print("  The link is active as long as this window is open.")
    print("  Press Ctrl+C to stop.")
    print("=" * 60)

    try:
        import time
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\nTunnel closed.")
        ngrok.disconnect(public_url)
        ngrok.kill()


# ── ENTRY POINT ───────────────────────────────────────────────────────────────
def _parse_port(args, default=5050):
    for arg in args:
        if "--port" in arg:
            try:
                return int(arg.split("=")[-1].strip())
            except ValueError:
                pass
    return default


if __name__ == "__main__":
    args = sys.argv[1:]
    mode = args[0] if args else "export"

    if mode == "serve":
        app, port = run_server(_parse_port(args))
        _start_flask(app, port)

    elif mode == "tunnel":
        run_tunnel(_parse_port(args))

    elif mode == "export":
        export_to_files()

    else:
        print(__doc__)
